import sqlite3
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.reader.excel import load_workbook
import os
from time import sleep


# Criando as views no banco de dados para pode construir os gráficos posteriormente
def creating_views():
    try:
        print(yellow_text('Criando as VIEWS no banco de dados...'))
        conn = sqlite3.connect('BOOKS.db')
        cur = conn.cursor()
        cur.execute('DROP VIEW IF EXISTS Prices;')
        cur.execute('''CREATE VIEW Prices AS
        SELECT Category, round(avg(Price), 2) AS "Average Price", 
        round(avg(Stars), 1) AS "Average Stars"
        FROM Books
        GROUP BY Category
        ORDER BY 3 DESC, 2;''')
        cur.execute('DROP VIEW IF EXISTS livros_por_categoria;')
        cur.execute('''CREATE VIEW livros_por_categoria AS
        SELECT Category, count(Book) AS "Nº de livros"
        FROM Books
        GROUP BY Category
        ORDER BY 2 DESC, 1;''')
        cur.execute('DROP VIEW IF EXISTS media_info_livros;')
        cur.execute('''CREATE VIEW media_info_livros AS
        SELECT 'Todos os Livros', round(avg(Price), 2) AS "Média Preços", 
        round(avg(Stars), 2) AS "Média Recomendações"
        FROM Books;''')
        conn.commit()
        conn.close()
        sleep(1.5)

    except:
        print(red_text('ERRO! As VIEWS não foram criadas corretamente.\n'
                       'Feche o banco de dados e tente novamente!'))
        exit()

    else:
        print(green_text('VIEWS criadas com sucesso!'))


# Gráfico de todas as categorias organizado por recomendação e preço
def chart_cats():
    # Nova conexão com o banco de dados para adicionar a VIEW "Prices" para fazer um gráfico no excel.
    try:
        conn = sqlite3.connect('BOOKS.db')
        cur = conn.cursor()
        cur.execute("SELECT * FROM PRICES")
        data = cur.fetchall()
        print(f'\n{blue_text("Dados de categoria, preço e recomendação:")} {data}')
        sleep(0.5)

        print(yellow_text('Criando tabela e gráfico no Excel com os dados acima...'))
        sleep(1)

        # Criando a planilha para montar a tabela que servirá de base para o gráfico
        wb = Workbook()
        ws_table = wb.create_sheet(title='Categorias')

        # Adicionando os nomes das colunas à planilha da tabela
        ws_table.append(('Categorias', 'Preço', 'Estrelas'))

        # Adicionando os dados armazenados no banco de dados à planilha da tabela
        for row in data:
            ws_table.append(row)

        # Fechando conexão com o banco de dados
        conn.close()

        # Criando o objeto gráfico do tipo barra em colunas
        chart1 = BarChart()
        chart1.type = "col"

        # Nomeando os eixos x e y
        chart1.y_axis.title = 'Preços'
        chart1.x_axis.title = 'Categorias'
        chart1.y_axis.majorGridlines = None

        # Adicionando as categorias ("cats" no eixo x) e os preços por categoria no eixo y
        data = Reference(ws_table, min_col=2, min_row=1, max_row=len(data) + 1)
        cats = Reference(ws_table, min_col=1, min_row=2, max_row=len(data) + 1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)

        # Posição da legenda no gráfico ("r" de "right", isto é, à direita do gráfico)
        chart1.legend.position = "r"

        # Altura e largura
        chart1.height = 12
        chart1.width = 35

        # Criando outro gráfico para adicionar a coluna estrelas ao eixo y do gráfico principal
        chart_stars = LineChart()
        data2 = Reference(ws_table, min_col=3, min_row=1, max_row=len(data) + 1)
        chart_stars.add_data(data2, titles_from_data=True)
        chart_stars.y_axis.axId = 20
        chart_stars.y_axis.title = "Estrelas"

        # Exibir o eixo y do segundo gráfico à direita, definindo-o para cruzar o eixo x em seu ponto máximo
        chart1.y_axis.crosses = "max"
        chart1 += chart_stars

        # Criando a planilha do gráfico e adicionando o gráfico montado a ela
        ws_chart = wb.create_sheet(title='graph categorias')
        ws_chart.add_chart(chart1, "A1")

        # Removendo a planilha vazia padrão inerente à criação do arquivo xlsx
        standard_sheet = wb['Sheet']
        wb.remove(standard_sheet)

        # Salvando as planilhas no formato xlsx (Excel)
        wb.save("livraria.xlsx")

        # Mensagem
        print(green_text('Planilha com tabela e gráfico de média de preço e recomendação'
                         ' por categoria criada com sucesso!'))
        sleep(1.5)

    except:
        print(red_text('ERRO! Dados da VIEW "PRICES" não foram salvos corretamente!\n'
                       'Feche a planilha e/ou o banco de dados e tente novamente!'))
        exit()


# Gráfico com as informações da média de todos os livros
def chart_livros():
    try:
        # Nova conexão com o banco de dados para adicionar a VIEW "Prices" para fazer um gráfico no excel.
        conn = sqlite3.connect('BOOKS.db')
        cur = conn.cursor()
        cur.execute("SELECT * FROM media_info_livros")
        data = cur.fetchall()
        print(f'\n{blue_text("Dados de média de preço e recomendação de todos os livros:")} {data}')
        sleep(0.5)

        print(yellow_text('Criando tabela e gráfico no Excel com os dados acima...'))
        sleep(1)

        # Criando a planilha para montar a tabela que eservirá de base para o gráfico
        wb = Workbook()
        wb = load_workbook('livraria.xlsx')
        ws_table = wb.create_sheet(title='Livros')

        # Adicionando os nomes das colunas à planilha da tabela
        ws_table.append(('Total de Livros', 'Média Preço', 'Média Recomendação'))

        # Adicionando os dados armazenados no banco de dados à planilha da tabela
        for row in data:
            ws_table.append(row)

        # Fechando conexão com o banco de dados
        conn.close()

        # Criando o objeto gráfico do tipo barra em colunas
        chart2 = BarChart()
        chart2.type = "col"

        # Nomeando o eixo y
        chart2.y_axis.title = 'Preços/Recomendações'

        # Escolhendo o estilo de cores (estilo 12 recomendado na própria documentação)
        chart2.style = 12

        # Adicionando as categorias ("cats") e os dados referentes a cada uma ("data") da tabela ao gráfico
        data = Reference(ws_table, min_col=2, min_row=1, max_row=len(data) + 1, max_col=3)
        books = Reference(ws_table, min_col=1, min_row=2, max_row=len(data) + 1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(books)

        # Posição da legenda no gráfico ("r" de "right", isto é, à direita do gráfico)
        chart2.legend.position = "r"

        # Altura e largura
        chart2.height = 12
        chart2.width = 35

        # Criando a planilha "graph livros" do gráfico e adicionando-o a ela.
        ws_chart = wb.create_sheet(title='graph livros')
        ws_chart.add_chart(chart2, "A1")

        # Salvando as planilhas no formato xlsx (Excel)
        wb.save("livraria.xlsx")

        # Mensagem
        print(green_text('Planilha com tabela e gráfico de média de preço e recomendação'
                         ' de todos os livros criada com sucesso!'))
        sleep(1.5)

    except:
        print(red_text('ERRO! Dados da VIEW "media_info_livros" não foram salvos corretamente!\n'
                       'Feche a planilha e/ou o banco de dados e tente novamente!'))
        exit()


# Gráfico com o número de livros por categoria
def chart_n_livros_por_categoria():
    try:
        # Nova conexão com o banco de dados para adicionar a VIEW "livros_por_categoria" para fazer um gráfico no excel.
        conn = sqlite3.connect('BOOKS.db')
        cur = conn.cursor()
        cur.execute("SELECT * FROM livros_por_categoria")
        data = cur.fetchall()
        print(f'\n{blue_text("Dados de quantidade de livros por categoria:")} {data}')
        sleep(0.5)

        print(yellow_text('Criando tabela e gráfico no Excel com os dados acima...'))
        sleep(1)

        # Criando a planilha para montar a tabela que servirá de base para o gráfico
        wb = Workbook()
        wb = load_workbook('livraria.xlsx')
        ws_table = wb.create_sheet(title='Livros por Categoria')

        # Adicionando os nomes das colunas à planilha da tabela
        ws_table.append(('Categorias', 'Nº de Livros'))

        # Adicionando os dados armazenados no banco de dados à planilha da tabela
        for row in data:
            ws_table.append(row)

        # Fechando conexão com o banco de dados
        conn.close()

        # Criando o objeto gráfico do tipo barra em colunas
        chart3 = BarChart()
        chart3.type = "col"

        # Nomeando os eixos x e y
        chart3.y_axis.title = 'Quantidade de livros'
        chart3.x_axis.title = 'Categorias'

        # Escolhendo o estilo de cores (estilo 12 recomendado na própria documentação)
        chart3.style = 12

        # Adicionando as categorias ("cats") e os dados referentes a cada uma ("data") da tabela ao gráfico
        data = Reference(ws_table, min_col=2, min_row=1, max_row=len(data) + 1)
        cats = Reference(ws_table, min_col=1, min_row=2, max_row=len(data) + 1)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)

        # Tirando a legenda
        chart3.legend = None

        # Altura e largura
        chart3.height = 12
        chart3.width = 35

        # Criando a planilha do gráfico e adicionando o gráfico montado a ela
        ws_chart = wb.create_sheet(title='graph Livros por Categoria')
        ws_chart.add_chart(chart3, "A1")

        # Salvando as planilhas no formato xlsx (Excel)
        wb.save("livraria.xlsx")

        # Mensagem
        print(green_text('Planilha com tabela e gráfico da quantidade de livros por categoria!'))
        sleep(1.5)

    except:
        print(red_text('ERRO! Dados da VIEW "livros_por_categoria" não foram salvos corretamente!\n'
                       'Feche a planilha e/ou o banco de dados e tente novamente!'))
        exit()


# Abrir planilha
def open_sheets(sheet='livraria.xlsx'):
    try:
        print(f'\n{yellow_text("Abrindo planilha: ")}{sheet} {yellow_text("...")}')
        sleep(0.5)
        os.system(f"start {sheet}")

    except:
        print(f'{red_text(sheet)} não existe')

    finally:
        print(blue_text('\nFim do processo!'))


# Texto em vermelho
def red_text(text):
    x = f'\033[1;31m{text}\033[m'
    return x


# Texto em verde
def green_text(text):
    x = f'\033[1;32m{text}\033[m'
    return x


# Texto em amarelo
def yellow_text(text):
    x = f'\033[1;33m{text}\033[m'
    return x


# Texto em azul
def blue_text(text):
    x = f'\033[1;34m{text}\033[m'
    return x




